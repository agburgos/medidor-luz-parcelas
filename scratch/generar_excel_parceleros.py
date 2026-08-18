import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

with open('/tmp/parcelas_data.json') as f:
    parcelas = json.load(f)

parcelas.sort(key=lambda p: p['numero'])

wb = Workbook()
ws = wb.active
ws.title = "Datos Parceleros"

FONT = "Arial"
azul = Font(name=FONT, size=11, bold=True, color="FFFFFF")
fill_header = PatternFill("solid", fgColor="1D4ED8")
fill_completar = PatternFill("solid", fgColor="FFFF00")
fill_ok = PatternFill("solid", fgColor="E8F5E9")
fill_leyenda = PatternFill("solid", fgColor="FEF9C3")
thin = Side(style="thin", color="D1D5DB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# --- Título y leyenda ---
ws.merge_cells("A1:F1")
ws["A1"] = "COPOSA — Datos para creación de cuentas de parceleros"
ws["A1"].font = Font(name=FONT, size=14, bold=True)

ws.merge_cells("A2:F2")
ws["A2"] = "Por favor completa SOLO las celdas amarillas (Teléfono y, si falta, Email). No modifiques el número de parcela ni el nombre."
ws["A2"].font = Font(name=FONT, size=10, italic=True)
ws["A2"].fill = fill_leyenda

ws.merge_cells("A3:F3")
ws["A3"] = "Ejemplo de fila ya completa: Parcela 1, Alex Balladares, alex@correo.cl, +56912345678"
ws["A3"].font = Font(name=FONT, size=9, italic=True, color="6B7280")

headers = ["Parcela #", "Nombre Dueño", "Email actual", "Email (completar si falta)", "Teléfono (completar)", "Notas"]
header_row = 5
for col, h in enumerate(headers, start=1):
    c = ws.cell(row=header_row, column=col, value=h)
    c.font = azul
    c.fill = fill_header
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border

row = header_row + 1
for p in parcelas:
    tiene_email = bool(p.get("email"))
    ws.cell(row=row, column=1, value=p["numero"]).border = border
    ws.cell(row=row, column=2, value=p["nombre_dueno"]).border = border
    c_email = ws.cell(row=row, column=3, value=p.get("email") or "")
    c_email.border = border
    c_email_nuevo = ws.cell(row=row, column=4, value="")
    c_email_nuevo.border = border
    c_tel = ws.cell(row=row, column=5, value="")
    c_tel.border = border
    ws.cell(row=row, column=6, value="").border = border

    if tiene_email:
        c_email.fill = fill_ok
    else:
        c_email_nuevo.fill = fill_completar
    c_tel.fill = fill_completar

    for col in range(1, 7):
        ws.cell(row=row, column=col).font = Font(name=FONT, size=10)
    row += 1

widths = [10, 26, 28, 30, 20, 30]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A6"

out_path = "/Users/agarridob/Proyectos/medidor-luz-parcelas/scratch/COPOSA_datos_parceleros.xlsx"
wb.save(out_path)
print("Guardado:", out_path, "| filas:", row - header_row - 1)
